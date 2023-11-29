import pymysql
from openpyxl import Workbook
from openpyxl import load_workbook


def insert_data(conn, cur, data):
    sql = 'insert into bookshelf (category, bookshelf_num, column_num, classification_number, author_symbol) values(%s, %s, %s, %s, %s)'  # SQL 문장 정의 (여기서는 %s만 사용 가능)

    iter_rows = iter(data)  # 아까 메모리에 저장했던 데이터들을 stream 값으로 변환
    next(iter_rows)  # 첫번째 row를 넘기는 목적으로 활용. 일반적으로 엑셀 첫 줄에 column 속성명을 입력하니까. 
    for row in iter_rows:
        cur.execute(sql, (row[0].value, row[1].value, row[2].value, row[3].value, row[4].value))  # 동적 SQL 문을 반복 실행. 파이썬이 알아서 해당 column에 적절한 타입으로 변환하여 대입함.
    conn.commit()  # DB 데이터 삽입 확정

def find_data():
    # DB 연결
    conn = pymysql.connect (host='localhost', port=3307, user='root', password='1234', 
                        db='ajoubooking', charset='utf8')

    try:
        with conn.cursor() as cur:  # DB와 상호작용하기 위해 연결해주는 cursor 객체
            wb = load_workbook("C:/Users/rlaeo/develop-world/Python-world/AJOU Booking/AJOU Booking.xlsx", data_only=True)  # 경로에 있는 xlsx 파일을 로드한다. (옵션은 재읽기가 가능하게 설정)
            ws = wb['시트1']  # 데이터를 가져올 시트를 선택
            find_all_data = ws.rows  # 시트의 모든 셀들을 순차적으로 읽어와서 튜플 형식으로 메모리에 저장
            insert_data(conn, cur, find_all_data)      
    
    finally:
        conn.close()  # DB 연결 종료
        wb.close()  # 엑셀 파일 닫기

# 현 모듈이 프로그램의 시작점일 때만 이 함수 실행
if __name__ == '__main__':
    find_data()